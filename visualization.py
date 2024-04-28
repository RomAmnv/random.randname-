# -*- coding: utf-8 -*-

import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import matplotlib.dates as mdates
import seaborn as sns

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from wordcloud import WordCloud
from nltk.tokenize import word_tokenize
import nltk

nltk.download('punkt')
nltk.download('stopwords')

import pymorphy3
from string import punctuation

# Загрузка данных
def make_excel(data):

    # Преобразование timestamp в формат даты
    data['timestamp'] = pd.to_datetime(data['timestamp'], format='mixed', errors='coerce')

    # предобработка
    # столбец question_2
    text = data['question_2'].astype(str).dropna()
    text = " ".join(text)
    words = word_tokenize(text)
    stop_words = set(nltk.corpus.stopwords.words('russian'))
    custom_stop_words = {'особенно', 'крайне', 'понравилось', 'очень', 'отличное'}
    filtered_words = [word for word in words if word.lower() not in stop_words and word.lower() not in custom_stop_words]
    filtered_text_2 = " ".join(filtered_words)


    # столбец question_3
    text = data['question_3'].astype(str).dropna()
    text = " ".join(text)
    words = word_tokenize(text)
    stop_words = set(nltk.corpus.stopwords.words('russian'))
    custom_stop_words = {'сложности', 'сложными', 'понять', 'понимания', 'сложной', 'это', 'некоторые', 'сложно', 'пониманием'}
    filtered_words = [word for word in words if word.lower() not in stop_words and word.lower() not in custom_stop_words]
    filtered_text_3 = " ".join(filtered_words)


    # столбец question_4
    text = data['question_4'].astype(str).dropna()
    text = " ".join(text)
    words = word_tokenize(text)
    stop_words = set(nltk.corpus.stopwords.words('russian'))
    custom_stop_words = {'предложить', 'добавить', 'ввести', 'организовать', 'каждой', 'провести', 'хотелось', 'устроить', 'предоставить'}
    filtered_words = [word for word in words if word.lower() not in stop_words and word.lower() not in custom_stop_words]
    filtered_text_4 = " ".join(filtered_words)

    # столбец question_5
    text = data['question_5'].astype(str).dropna()
    text = " ".join(text)
    words = word_tokenize(text)
    stop_words = set(nltk.corpus.stopwords.words('russian'))
    custom_stop_words = {'применение', 'работа'}
    filtered_words = [word for word in words if word.lower() not in stop_words and word.lower() not in custom_stop_words]
    filtered_text_5 = " ".join(filtered_words)

    # функция предобработки текстов
    morph = pymorphy3.MorphAnalyzer()
    punktuations = list(punctuation)
    punktuations.append('—')

    def preproc (text):
        token = word_tokenize(text.lower())
        words_without_punkt = [i for i in token if (i not in punktuations)]
        lem = [morph.parse(i)[0].normal_form for i in words_without_punkt]
        return lem

    # обычные графики
    def survey_statistics(data):
        # График числа участвовавших студентов в опросе после конкретного вебинара (по датам)
        plt.figure(figsize=(7, 4))
        students_per_webinar = data['timestamp'].dt.date.value_counts().sort_index()
        students_per_webinar.plot(kind='bar', color=(0/255, 254/255, 8/255))  # RGB(0, 254, 8)
        plt.title('Число участвовавших студентов в опросе после вебинаров', fontweight='bold')
        plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%Y-%m-%d'))
        plt.xlabel('Дата')
        plt.ylabel('Число студентов')
        plt.tight_layout()
        plt.savefig('students_per_webinar.png')
        plt.close()


    def relevance_of_reviews(data):
        # Круговая диаграмма по соотношению релевантности отзывов к вебинарам
        relevant_counts = data['is_relevant'].value_counts()
        plt.figure(figsize=(6, 4))
        green_color = (0/255, 254/255, 8/255, 1)  # RGB(0, 254, 8)
        dark_green_color = (3/255, 25/255, 17/255, 1)  # RGB(3, 25, 17)
        plt.pie(relevant_counts, labels=['Релевантный', 'Нерелевантный'], autopct='%1.1f%%', colors=[green_color, dark_green_color])
        plt.title('Соотношение релевантности отзывов к вебинарам', fontweight='bold')
        plt.tight_layout()
        plt.savefig('relevant_feedback_pie.png')
        plt.close()


    def positive_or_negative(data):
        # Круговая диаграмма по соотношению положительных/отрицательных отзывов
        positive_counts = data['is_positive'].value_counts()
        plt.figure(figsize=(6, 4))
        green_color = (0/255, 254/255, 8/255, 1)  # RGB(0, 254, 8)
        dark_green_color = (3/255, 25/255, 17/255, 1)  # RGB(3, 25, 17)
        plt.pie(positive_counts, labels=['Позитивный', 'Негативный'], autopct='%1.1f%%', colors=[green_color, dark_green_color])
        plt.title('Соотношение положительных/отрицательных отзывов', fontweight='bold')
        plt.tight_layout()
        plt.savefig('positive_negative_feedback_pie.png')
        plt.close()


    def distribution_of_response_lengths(data):
        # Распределение длины комментариев по каждому вопросу (Рейтинг)
        plt.figure(figsize=(7, 4))
        for i in range(1, 6):
            plt.hist(data[f'question_{i}'].str.len(), bins=20, alpha=0.5, label=f'Вопрос {i}')
        plt.title('Распределение длины комментариев по каждому вопросу', fontweight='bold')
        plt.xlabel('Длина комментария')
        plt.ylabel('Число комментариев')
        plt.legend()
        plt.tight_layout()
        plt.savefig('comment_length_distribution.png')
        plt.close()


    def distribution_of_response_lengths(data):
        # Распределение длины комментариев по каждому вопросу (Рейтинг)
        lengths = data[[f'question_{i}' for i in range(1, 6)]].applymap(lambda x: len(str(x)))
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.boxplot(lengths.values, labels=lengths.columns, showfliers=False)
        ax.set_title('Распределение длины комментариев по каждому вопросу', fontweight='bold')
        ax.set_xlabel('Вопрос')
        ax.set_ylabel('Длина комментария')
        plt.xticks(rotation=45)
        plt.tight_layout()
        plt.savefig('comment_length_distribution.png')
        plt.close()


    def negative_reviews_of_the_objects(data):
        # Сопоставление числовых значений с нужными метками
        labels = {0: 'Вебинар', 1: 'Программа', 2: 'Преподаватель'}
        # Замена числовых значений на нужные метки
        data['object'] = data['object'].replace(labels)
        # Выборка данных по негативным отзывам
        negative_counts_per_webinar = data[data['is_positive'] == False]['object'].value_counts()
        # Построение столбчатой диаграммы
        plt.figure(figsize=(5, 3))
        negative_counts_per_webinar.plot(kind='bar', color=(0/255, 254/255, 8/255))  # RGB(0, 254, 8)
        plt.title('Негативные отзывы по объектам', fontweight='bold')
        plt.xlabel('Объект')
        plt.ylabel('Число негативных отзывов')
        plt.xticks(rotation=45, ha='right')
        plt.grid(axis='y')
        plt.tight_layout()
        plt.savefig('negative_feedback_per_webinar.png')
        plt.close()

    # облака слов
    def create_wordcloud_2(text, width=400, height=400, background_color='white', filename='wordcloud_2.png'):
        # облако слов для question_2 Что вам больше всего понравилось в теме вебинара и почему?
        wordcloud = WordCloud(width=width, height=height, background_color=background_color)
        wordcloud.generate(text)
        plt.figure(figsize=(10, 4))
        plt.axis('off')
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.title('Что вам больше всего понравилось в теме вебинара и почему?', fontweight='bold')
        plt.tight_layout()
        plt.savefig(filename)
        plt.close()


    def create_wordcloud_3(text, width=400, height=400, background_color='white', filename='wordcloud_3.png'):
        # облако слов для question_3 Были ли моменты в вебинаре, которые вызвали затруднения в понимании материала? Можете описать их?
        wordcloud = WordCloud(width=width, height=height, background_color=background_color)
        wordcloud.generate(text)
        plt.figure(figsize=(10, 4))
        plt.axis('off')
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.title('Были ли моменты в вебинаре, которые вызвали затруднения в понимании материала?', fontweight='bold')
        plt.tight_layout()
        plt.savefig(filename)
        plt.close()


    def create_wordcloud_4(text, width=400, height=400, background_color='white', filename='wordcloud_4.png'):
        # облако слов для question_4 Какие аспекты вебинара, по вашему мнению, нуждаются в улучшении и какие конкретные изменения вы бы предложили?
        wordcloud = WordCloud(width=width, height=height, background_color=background_color)
        wordcloud.generate(text)
        plt.figure(figsize=(10, 4))
        plt.axis('off')
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.title('Какие аспекты вебинара нуждаются в улучшении?', fontweight='bold')
        plt.tight_layout()
        plt.savefig(filename)
        plt.close()


    def create_wordcloud_5(text, width=400, height=400, background_color='white', filename='wordcloud_5.png'):
        # облако слов для question_5 Есть ли темы или вопросы, которые вы бы хотели изучить более подробно в следующих занятиях?
        wordcloud = WordCloud(width=width, height=height, background_color=background_color)
        wordcloud.generate(text)
        plt.figure(figsize=(10, 4))
        plt.axis('off')
        plt.imshow(wordcloud, interpolation='bilinear')
        plt.title('Вопросы, которые вы бы хотели изучить более подробно?', fontweight='bold')
        plt.tight_layout()
        plt.savefig(filename)
        plt.close()

    def plot_top_bigrams_2(text, width=400, height=400, background_color='white', filename='plot_top_bigrams_2.png'):
        bigrams_series = (pd.Series(nltk.ngrams(preproc(filtered_text_2), 2)).value_counts())[:10]
        bigrams_top = pd.DataFrame(bigrams_series.sort_values(ascending=False))
        bigrams_top = bigrams_top.reset_index().rename(columns={'index': 'bigrams', 'count':'counts'})
        bigrams_top['bigrams'] = bigrams_top['bigrams'].astype(str)
        plt.figure(figsize=(6,4))
        sns.catplot(x = 'counts' , y='bigrams', kind="bar", palette="vlag", data=bigrams_top, height=8.27, aspect=11.7/8.27)
        plt.title('TOP 10 bigram')
        plt.tight_layout()
        plt.savefig(filename)
        plt.close()

    def plot_top_bigrams_3(text, width=400, height=400, background_color='white', filename='plot_top_bigrams_3.png'):
        bigrams_series = (pd.Series(nltk.ngrams(preproc(filtered_text_3), 2)).value_counts())[:10]
        bigrams_top = pd.DataFrame(bigrams_series.sort_values(ascending=False))
        bigrams_top = bigrams_top.reset_index().rename(columns={'index': 'bigrams', 'count':'counts'})
        bigrams_top['bigrams'] = bigrams_top['bigrams'].astype(str)
        plt.figure(figsize=(6,4))
        sns.catplot(x = 'counts' , y='bigrams', kind="bar", palette="vlag", data=bigrams_top, height=8.27, aspect=11.7/8.27)
        plt.title('TOP 10 bigram')
        plt.tight_layout()
        plt.savefig(filename)
        plt.close()


    def plot_top_bigrams_4(text, width=400, height=400, background_color='white', filename='plot_top_bigrams_4.png'):
        bigrams_series = (pd.Series(nltk.ngrams(preproc(filtered_text_4), 2)).value_counts())[:10]
        bigrams_top = pd.DataFrame(bigrams_series.sort_values(ascending=False))
        bigrams_top = bigrams_top.reset_index().rename(columns={'index': 'bigrams', 'count':'counts'})
        bigrams_top['bigrams'] = bigrams_top['bigrams'].astype(str)
        plt.figure(figsize=(6,4))
        sns.catplot(x = 'counts' , y='bigrams', kind="bar", palette="vlag", data=bigrams_top, height=8.27, aspect=11.7/8.27)
        plt.title('TOP 10 bigram')
        plt.tight_layout()
        plt.savefig(filename)
        plt.close()


    def plot_top_bigrams_5(text, width=400, height=400, background_color='white', filename='plot_top_bigrams_5.png'):
        bigrams_series = (pd.Series(nltk.ngrams(preproc(filtered_text_5), 2)).value_counts())[:10]
        bigrams_top = pd.DataFrame(bigrams_series.sort_values(ascending=False))
        bigrams_top = bigrams_top.reset_index().rename(columns={'index': 'bigrams', 'count':'counts'})
        bigrams_top['bigrams'] = bigrams_top['bigrams'].astype(str)
        plt.figure(figsize=(6,4))
        sns.catplot(x = 'counts' , y='bigrams', kind="bar", palette="vlag", data=bigrams_top, height=8.27, aspect=11.7/8.27)
        plt.title('TOP 10 bigram')
        plt.tight_layout()
        plt.savefig(filename)
        plt.close()

    # вызов функций
    students_per_webinar_data = survey_statistics(data)
    relevance_data = relevance_of_reviews(data)
    positive_negative_data = positive_or_negative(data)
    relevant_feedback = distribution_of_response_lengths(data)
    negative_reviews_data = negative_reviews_of_the_objects(data)

    create_wordcloud_2 = create_wordcloud_2(filtered_text_2)
    create_wordcloud_3 = create_wordcloud_3(filtered_text_3)
    create_wordcloud_4 = create_wordcloud_4(filtered_text_4)
    create_wordcloud_5 = create_wordcloud_5(filtered_text_5)

    plot_top_bigrams_2 = plot_top_bigrams_2(filtered_text_2)
    plot_top_bigrams_3 = plot_top_bigrams_3(filtered_text_3)
    plot_top_bigrams_4 = plot_top_bigrams_4(filtered_text_4)
    plot_top_bigrams_5 = plot_top_bigrams_5(filtered_text_5)

    # Создание Excel-файла
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Исходные данные"

    # Записываем данные на первый лист
    for row in dataframe_to_rows(data, index=False, header=True):
        worksheet.append(row)

    # Автоматически подогнать ширину столбцов под содержимое
    for col in range(worksheet.max_column):
        column_letter = get_column_letter(col + 1)  # Получить букву столбца
        worksheet.column_dimensions[column_letter].auto_size = True

    worksheet1 = workbook.create_sheet('Участие в опросе')
    worksheet2 = workbook.create_sheet('Релевантность отзывов')
    worksheet3 = workbook.create_sheet('Окраска отзывов')
    worksheet4 = workbook.create_sheet('Длина комментариев')
    worksheet5 = workbook.create_sheet('Объект комментариев')
    worksheet6 = workbook.create_sheet('Облака слов и биграмы')

    # Добавление таблиц и изображений на листы
    # Лист 1
    # Таблица
    students_per_webinar_data = data['timestamp'].dt.date.value_counts().reset_index()
    students_per_webinar_data.columns = ['Дата', 'Число студентов']
    worksheet1.append(['Дата', 'Число студентов'])
    for row in dataframe_to_rows(students_per_webinar_data, index=False, header=False):
        worksheet1.append(row)

    # Автоматически подогнать ширину столбцов под содержимое
    for col in range(worksheet1.max_column):
        column_letter = get_column_letter(col + 1)  # Получить букву столбца
        worksheet1.column_dimensions[column_letter].auto_size = True

    # Изображение
    img_positive = Image('students_per_webinar.png')
    img_positive_anchor = worksheet1.add_image(img_positive, 'D1')


    # Лист 2
    # Таблица
    relevant_data = data['is_relevant'].value_counts().reset_index()
    relevant_data.columns = ['Релевантность', 'Количество']
    worksheet2.append(['Релевантность', 'Количество'])  # Добавляем заголовок таблицы
    for row in dataframe_to_rows(relevant_data, index=False, header=False):
        worksheet2.append(row)

    # Автоматически подогнать ширину столбцов под содержимое
    for col in range(worksheet2.max_column):
        column_letter = get_column_letter(col + 1)  # Получить букву столбца
        worksheet2.column_dimensions[column_letter].auto_size = True  # Используем worksheet2

    # Изображение
    img_relevant = Image('relevant_feedback_pie.png')
    worksheet2.add_image(img_relevant, 'D1')


    # Лист 3
    # Таблица
    positive_data = data['is_positive'].value_counts().reset_index()
    positive_data.columns = ['Тональность', 'Количество']
    for row in dataframe_to_rows(positive_data, index=False, header=True):
        worksheet3.append(row)

    # Автоматически подогнать ширину столбцов под содержимое
    for col in range(worksheet3.max_column):
        column_letter = get_column_letter(col + 1)  # Получить букву столбца
        worksheet3.column_dimensions[column_letter].auto_size = True

    # Изображение
    img_positive = Image('positive_negative_feedback_pie.png')
    worksheet3.add_image(img_positive, 'D1')


    # Лист 4
    # Таблица
    for i in range(1, 6):
        question_data = data[f'question_{i}'].str.len().value_counts().reset_index()
        question_data.columns = [f'Длина комментария (Вопрос {i})', 'Количество']
        for row in dataframe_to_rows(question_data, index=False, header=True):
            worksheet4.append(row)
        worksheet4.append([])  # Добавляем пустую строку для разделения вопросов

    # Автоматически подогнать ширину столбцов под содержимое
    for col in range(worksheet4.max_column):
        column_letter = get_column_letter(col + 1)  # Получить букву столбца
        worksheet4.column_dimensions[column_letter].auto_size = True

    # Изображение
    img_comment_length = Image('comment_length_distribution.png')
    worksheet4.add_image(img_comment_length, 'D1')


    # Лист 5
    # Таблица
    negative_counts_data = data[data['is_positive'] == False]['object'].value_counts().reset_index()
    negative_counts_data.columns = ['Объект', 'Число негативных отзывов']
    for row in dataframe_to_rows(negative_counts_data, index=False, header=True):
        worksheet5.append(row)

    # Автоматически подогнать ширину столбцов под содержимое
    for col in range(worksheet5.max_column):
        column_letter = get_column_letter(col + 1)  # Получить букву столбца
        worksheet5.column_dimensions[column_letter].auto_size = True

    # Изображение
    img_negative_feedback = Image('negative_feedback_per_webinar.png')
    worksheet5.add_image(img_negative_feedback, 'D1')


    # Лист 6
    # Облака слов
    wordcloud_2 = Image('wordcloud_2.png')
    worksheet6.add_image(wordcloud_2, 'A1')

    wordcloud_3 = Image('wordcloud_3.png')
    worksheet6.add_image(wordcloud_3, 'A22')

    wordcloud_4 = Image('wordcloud_4.png')
    worksheet6.add_image(wordcloud_4, 'A44')

    wordcloud_5 = Image('wordcloud_5.png')
    worksheet6.add_image(wordcloud_5, 'A64')

    # Биграммы
    top_bigrams_2 = Image('plot_top_bigrams_2.png')
    worksheet6.add_image(top_bigrams_2, 'Q1')

    top_bigrams_3 = Image('plot_top_bigrams_3.png')
    worksheet6.add_image(top_bigrams_3, 'Q22')

    top_bigrams_4 = Image('plot_top_bigrams_4.png')
    worksheet6.add_image(top_bigrams_4, 'Q44')

    top_bigrams_5 = Image('plot_top_bigrams_5.png')
    worksheet6.add_image(top_bigrams_5, 'Q64')


    # Сохранение файла
    workbook.save('results.xlsx')
    return workbook